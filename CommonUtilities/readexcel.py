import openpyxl
import os


def read_data_from_excel(file_location, sheet_name, user_name):
    wk = openpyxl.load_workbook(file_location)
    sh = wk[sheet_name]
    rows = sh.max_row
    columns = sh.max_column
    print(wk.sheetnames)
    print("Active sheet is: " + wk.active.title)
    print("Current Directory: " + os.getcwd())
    product_detail_list = list()
    for i in range(1, rows+1):
        # for j in range(1, columns+1):
        for j in range(2, 3):
            l1 = list()
            user = sh.cell(i, j)
            p_title = sh.cell(i, j+1)
            p_desc = sh.cell(i, j+2)
            p_image = sh.cell(i, j+3)
            if user.value == user_name and p_title.value is not None and p_desc.value is not None:
                l1.append(p_title.value[0:150])
                l1.append(p_desc.value)
                l1.append(p_image.value)
            else:
                break
        if l1 != []:
            product_detail_list.append(l1)
    return product_detail_list

#print(read_data_from_excel("/Users/amittiwari/Documents/TGM/TestData/TGM-Auto.xlsx", "Post Data", "aarfa_kurties_manufacturer_"))